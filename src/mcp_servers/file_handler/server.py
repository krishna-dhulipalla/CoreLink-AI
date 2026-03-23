"""File Handler MCP server.

Provides generic document retrieval and parsing for evaluator-supplied files.
"""

from __future__ import annotations

import csv
import io
import json
import re
from typing import Optional

import httpx
from mcp.server.fastmcp import FastMCP

mcp = FastMCP("File Handler")

DEFAULT_TIMEOUT = 30.0
MAX_TEXT_CHARS = 12_000
_SEARCH_STOP_WORDS = {
    "the",
    "and",
    "for",
    "with",
    "that",
    "this",
    "from",
    "what",
    "using",
    "were",
    "was",
    "into",
    "year",
    "fiscal",
    "calendar",
    "total",
}


def _is_probably_binary(raw: bytes) -> bool:
    if not raw:
        return False
    sample = raw[:2048]
    if b"\x00" in sample:
        return True
    non_text = sum(
        1
        for byte in sample
        if byte < 9 or (13 < byte < 32) or byte > 126
    )
    return non_text / max(1, len(sample)) > 0.30


def _magic_format(raw: bytes, url: str, content_type: str) -> str:
    sample = raw[:16]
    lowered_url = url.lower().split("?")[0]
    lowered_ct = (content_type or "").lower()
    if sample.startswith(b"%PDF-"):
        return "pdf"
    if sample.startswith((b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")):
        if lowered_url.endswith((".xlsx", ".xlsm")) or "spreadsheet" in lowered_ct:
            return "excel"
        if lowered_url.endswith((".docx", ".docm")) or "wordprocessing" in lowered_ct:
            return "word"
        return "archive"
    return ""


def _sniff_format(url: str, content_type: str, raw: bytes | None = None) -> str:
    """Infer file format from URL extension, content type, or magic bytes."""
    if raw:
        magic = _magic_format(raw, url, content_type)
        if magic:
            return magic

    url_lower = url.lower().split("?")[0]
    for ext, fmt in [
        (".pdf", "pdf"),
        (".xlsx", "excel"),
        (".xls", "excel"),
        (".docx", "word"),
        (".doc", "word"),
        (".csv", "csv"),
        (".json", "json"),
        (".png", "image"),
        (".jpg", "image"),
        (".jpeg", "image"),
        (".wav", "audio"),
        (".mp3", "audio"),
        (".mp4", "video"),
        (".avi", "video"),
        (".zip", "archive"),
        (".tar.gz", "archive"),
        (".rar", "archive"),
        (".txt", "text"),
        (".md", "text"),
    ]:
        if url_lower.endswith(ext):
            return fmt

    ct = (content_type or "").lower()
    if "pdf" in ct:
        return "pdf"
    if "spreadsheet" in ct:
        return "excel"
    if "wordprocessing" in ct:
        return "word"
    if "csv" in ct:
        return "csv"
    if "json" in ct:
        return "json"
    if "image" in ct:
        return "image"
    if "audio" in ct:
        return "audio"
    if "video" in ct:
        return "video"
    if "zip" in ct or "tar" in ct:
        return "archive"
    return "text"


def _format_header(url: str, fmt: str, size_kb: float, *, status: str = "OK", error: str = "") -> str:
    header = (
        f"FILE: {url.split('/')[-1].split('?')[0]}\n"
        f"FORMAT: {fmt.upper()} | SIZE: {size_kb:.1f} KB\n"
        f"STATUS: {status}\n"
    )
    if error:
        header += f"ERROR: {error}\n"
    return header + f"{'-' * 50}\n"


def _search_tokens(text: str) -> list[str]:
    return [
        token
        for token in re.findall(r"[a-z0-9]+", (text or "").lower())
        if len(token) > 1 and token not in _SEARCH_STOP_WORDS
    ]


def _parse_pdf(raw: bytes, page_start: int, page_limit: int, search_hint: str | None = None) -> tuple[str, int, int]:
    try:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(raw))
        total = len(reader.pages)
        actual_start = max(0, page_start)
        hint_tokens = set(_search_tokens(search_hint or ""))
        page_texts: list[str] = []
        if page_start == 0 and hint_tokens:
            best_score = 0
            best_index = 0
            for index, page in enumerate(reader.pages):
                text = page.extract_text() or ""
                page_texts.append(text)
                score = len(hint_tokens.intersection(_search_tokens(text)))
                if score > best_score:
                    best_score = score
                    best_index = index
            if best_score >= 2:
                actual_start = best_index
        end = min(actual_start + page_limit, total)
        pages = []
        for index in range(actual_start, end):
            text = page_texts[index] if index < len(page_texts) else (reader.pages[index].extract_text() or "")
            pages.append(f"[Page {index + 1}]\n{text.strip()}")
        return "\n\n".join(pages), total, actual_start
    except ImportError:
        return "[ERROR: pypdf not installed. Run: uv add pypdf]", 0, 0
    except Exception as exc:
        return f"[PDF parse error: {exc}]", 0, 0


def _parse_excel(raw: bytes, sheet: Optional[str], row_offset: int, row_limit: int) -> tuple[str, int]:
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        worksheet = workbook[sheet] if sheet and sheet in workbook.sheetnames else workbook.active
        all_rows = list(worksheet.iter_rows(values_only=True))
        total = len(all_rows)
        rows = all_rows[row_offset : row_offset + row_limit]
        lines = []
        for row in rows:
            lines.append("\t".join("" if value is None else str(value) for value in row))
        return "\n".join(lines), total
    except ImportError:
        return "[ERROR: openpyxl not installed. Run: uv add openpyxl]", 0
    except Exception as exc:
        return f"[Excel parse error: {exc}]", 0


def _parse_word(raw: bytes) -> str:
    try:
        from docx import Document

        document = Document(io.BytesIO(raw))
        return "\n".join(p.text for p in document.paragraphs if p.text.strip())
    except ImportError:
        return "[ERROR: python-docx not installed. Run: uv add python-docx]"
    except Exception as exc:
        return f"[Word parse error: {exc}]"


def _parse_csv(raw: bytes, row_offset: int, row_limit: int) -> tuple[str, int]:
    try:
        text = raw.decode("utf-8", errors="replace")
        try:
            all_rows = list(csv.reader(text.splitlines()))
        except csv.Error:
            all_rows = []
            for line in text.splitlines():
                if not line.strip():
                    continue
                try:
                    parsed = next(csv.reader([line]))
                except csv.Error:
                    parsed = [part.strip() for part in line.split(",")]
                all_rows.append(parsed)
        total = len(all_rows)
        rows = all_rows[row_offset : row_offset + row_limit]
        return "\n".join(",".join(row) for row in rows), total
    except Exception as exc:
        return f"[CSV parse error: {exc}]", 0


@mcp.tool()
def fetch_reference_file(
    url: str,
    page_start: int = 0,
    page_limit: int = 5,
    row_offset: int = 0,
    row_limit: int = 200,
    sheet: Optional[str] = None,
    format_hint: Optional[str] = None,
    search_hint: Optional[str] = None,
) -> str:
    """Download and parse a reference file from an evaluator-provided URL."""
    try:
        with httpx.Client(timeout=DEFAULT_TIMEOUT, follow_redirects=True) as client:
            response = client.get(url)
            response.raise_for_status()

        raw = response.content
        content_type = response.headers.get("content-type", "")
        size_kb = len(raw) / 1024
        fmt = format_hint.lower() if format_hint else _sniff_format(url, content_type, raw)
        header = _format_header(url, fmt, size_kb)

        if fmt == "pdf":
            content, total, actual_start = _parse_pdf(raw, page_start, page_limit, search_hint)
            if content.startswith("[PDF parse error:") or content.startswith("[ERROR:"):
                return _format_header(url, fmt, size_kb, status="PARSE_ERROR", error=content) + content[:MAX_TEXT_CHARS]
            meta = f"[Pages {actual_start + 1}-{min(actual_start + page_limit, total)} of {total}]\n"
            return header + meta + content[:MAX_TEXT_CHARS]

        if fmt == "excel":
            content, total = _parse_excel(raw, sheet, row_offset, row_limit)
            if content.startswith("[Excel parse error:") or content.startswith("[ERROR:"):
                return _format_header(url, fmt, size_kb, status="PARSE_ERROR", error=content) + content[:MAX_TEXT_CHARS]
            meta = f"[Rows {row_offset}-{row_offset + row_limit} of ~{total}]\n"
            return header + meta + content[:MAX_TEXT_CHARS]

        if fmt == "csv":
            content, total = _parse_csv(raw, row_offset, row_limit)
            if content.startswith("[CSV parse error:"):
                return _format_header(url, fmt, size_kb, status="PARSE_ERROR", error=content) + content[:MAX_TEXT_CHARS]
            meta = f"[Rows {row_offset}-{row_offset + row_limit} of ~{total}]\n"
            return header + meta + content[:MAX_TEXT_CHARS]

        if fmt == "word":
            content = _parse_word(raw)
            if content.startswith("[Word parse error:") or content.startswith("[ERROR:"):
                return _format_header(url, fmt, size_kb, status="PARSE_ERROR", error=content) + content[:MAX_TEXT_CHARS]
            return header + content[:MAX_TEXT_CHARS]

        if fmt == "json":
            try:
                parsed = json.loads(raw)
                pretty = json.dumps(parsed, indent=2)
                truncated = pretty[:MAX_TEXT_CHARS]
                suffix = "\n... [truncated, use pagination if needed]" if len(pretty) > MAX_TEXT_CHARS else ""
                return header + truncated + suffix
            except json.JSONDecodeError as exc:
                error = f"[JSON parse error: {exc}]"
                return _format_header(url, fmt, size_kb, status="PARSE_ERROR", error=error) + error

        if fmt == "image":
            return (
                _format_header(url, fmt, size_kb, status="UNSUPPORTED_FORMAT")
                + "[IMAGE FILE DETECTED]\n"
                + "This tool cannot extract pixel data from images. "
                + "If the image contains a chart or table, ask the evaluator to provide data in CSV or JSON format instead."
            )

        if fmt in {"audio", "video"}:
            return (
                _format_header(url, fmt, size_kb, status="UNSUPPORTED_FORMAT")
                + f"[{fmt.upper()} FILE DETECTED]\n"
                + "This tool cannot extract media streams. "
                + "If the task requires processing media files natively, an external specialized tool must be provided."
            )

        if fmt == "archive":
            return (
                _format_header(url, fmt, size_kb, status="UNSUPPORTED_FORMAT")
                + "[ARCHIVE FILE DETECTED]\n"
                + "This tool cannot extract or traverse archives directly."
            )

        if _is_probably_binary(raw):
            error = "Binary payload was detected but could not be parsed as PDF, Office, CSV, JSON, or text."
            return _format_header(url, "binary", size_kb, status="GARBLED_BINARY", error=error) + error

        text = raw.decode("utf-8", errors="replace")
        truncated = text[:MAX_TEXT_CHARS]
        suffix = "\n... [truncated]" if len(text) > MAX_TEXT_CHARS else ""
        return header + truncated + suffix

    except httpx.HTTPStatusError as exc:
        return f"[HTTP {exc.response.status_code}] Failed to fetch {url}: {exc}"
    except httpx.RequestError as exc:
        return f"[Network error] Could not reach {url}: {exc}"
    except Exception as exc:
        return f"[Unexpected error reading {url}]: {exc}"


@mcp.tool()
def list_reference_files(prompt_text: str) -> str:
    """Extract reference file URLs from the task prompt text."""
    urls: list[str] = []
    url_pattern = re.compile(r"https?://[^\s\)\]\"\'\,]+")
    found = url_pattern.findall(prompt_text)

    file_extensions = {".pdf", ".csv", ".xlsx", ".xls", ".docx", ".doc", ".json", ".txt", ".md"}
    for candidate in found:
        clean = candidate.rstrip(".,;)")
        if any(clean.lower().endswith(ext) for ext in file_extensions) or "download" in clean.lower() or "file" in clean.lower():
            urls.append(clean)
        elif clean not in urls:
            urls.append(clean)

    has_ref_signal = "REFERENCE FILES AVAILABLE" in prompt_text.upper() or "REFERENCE FILE" in prompt_text.upper()
    if not urls:
        suffix = ""
        if has_ref_signal:
            suffix = " Warning: 'REFERENCE FILES AVAILABLE' was present but no URLs were extracted."
        return "No reference file URLs detected in the prompt." + suffix

    lines = [
        "REFERENCE FILES DETECTED:" if has_ref_signal else "URLs found in prompt:",
        "",
    ]
    for index, candidate in enumerate(urls, start=1):
        ext = next(
            (
                ext
                for ext in [".pdf", ".csv", ".xlsx", ".xls", ".docx", ".json", ".txt"]
                if candidate.lower().endswith(ext)
            ),
            "unknown",
        )
        label = ext.lstrip(".").upper() if ext != "unknown" else "URL"
        lines.append(f"  {index}. [{label}] {candidate}")

    lines.append("")
    lines.append("Call fetch_reference_file(url=...) for each URL you need to read.")
    return "\n".join(lines)


if __name__ == "__main__":
    mcp.run()
